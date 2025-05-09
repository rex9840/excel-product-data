import os 
import json
from typing import Any
import pandas as pd
import openpyxl
import logging
import time
from .serializers import ProductItemSerializer
from . import models


logger = logging.getLogger() 


MAX_CHUNKS_SIZE = 100
FOREIGN_KEYS=["item_group_id","brand","gender","google_product_category","product_type","material","pattern","color"]  



class NullValueException(Exception): 
    def __init__(self, message="Null value found in the cell"):
        self.message = message
        super().__init__(self.message) 


def extract_header(file_path:str)->list[str]: 
    df = pd.read_excel(file_path,nrows=0)    
    headers = df.columns.to_list()
    return headers 


def map_workbook_Json(workbook:Any,start_row:int,stop_row:int,headers:list[str],start_time:float)->str:
    worksheet = workbook.active 
    data = []
    null_counter=0 
    for row in worksheet.iter_rows(min_row=start_row+1,max_row=stop_row):
        row_data = {}
        for index,cell in enumerate(row[:len(headers)]):  
            if cell.value is None:
                e =  NullValueException(f"Null value found in the cell {cell.coordinate}") 
                logger.warning(e.message)
                models.Log.objects.create(
                    message=e.__str__(), 
                    status = models.LogStatus.WARNING,
                    remarks = f"WARNING_{start_time}"
                ) 
                cell.value = ""
                null_counter += 1 
            if headers[index] == "shipping(country:price)":
                headers[index] = "shipping_cost"
            if headers[index] in FOREIGN_KEYS:
                row_data[headers[index].lower()] = {
                    "name": cell.value
                } 
            else: 
                row_data[headers[index].lower()] = cell.value         
                continue 
        if null_counter > len(headers):
             del data[-1]
             break
        data.extend([row_data])
    data = json.dumps(data)
    return data

# def serialize_and_save_json(filepath:str)->None:
#     start_time = time.time() 
#     models.Log.objects.create(
#             message = start_time,
#             status = models.LogStatus.INFO,
#             remarks = "START_TIME"
#     )
#
#     headers = extract_header(filepath)
#     workbook = openpyxl.load_workbook(filepath) 
#     worksheets = workbook.active 
#     max_row = worksheets.max_row
#     row_counter = max_row
#     start_row = 1
#     while(row_counter>0 or start_row<row_counter):  
#             try: 
#                 if row_counter < MAX_CHUNKS_SIZE:
#                     if start_row > row_counter and row_counter < MAX_CHUNKS_SIZE:
#                         end_row = max_row 
#                     else: 
#                         end_row = row_counter 
#                     data = map_workbook_Json(workbook,start_row,end_row,headers,start_time)
#                 else: 
#                     end_row = MAX_CHUNKS_SIZE + start_row
#                     data = map_workbook_Json(workbook,start_row,end_row,headers,start_time)
#                     start_row = end_row 
#                 row_counter = row_counter - MAX_CHUNKS_SIZE 
#                 data = json.loads(data) 
#                 serializer = ProductItemSerializer(data=data,many=True)         
#                 serializer.is_valid(raise_exception=True) 
#                 serializer.save() 
#                 
#             except Exception as e:
#                 logger.error(e.__str__()) 
#                 models.Log.objects.create(
#                     message=e.__str__(), 
#                     status = models.LogStatus.ERROR,
#                     remarks = f"ERROR_{start_time}" 
#                 ) 
#                 continue 
#     end_time = time.time() 
#     models.Log.objects.create(
#             message = end_time,
#             status = models.LogStatus.INFO,
#             remarks = f"END_TIME_{start_time}" 
#     ) 
#     rows_count = models.Log.objects.filter(remarks=f"ITEMS_{start_time}").count()
#     print(f"Time taken to process the file: {end_time - start_time} seconds")  
#     models.Log.objects.create(
#             message = f"{end_time - start_time}",
#             status = models.LogStatus.INFO,
#             remarks=f"TIME_TAKEN_{start_time}" 
#     )
#     os.remove(filepath) 
