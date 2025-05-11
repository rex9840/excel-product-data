import os
import json
import openpyxl
import logging
import time
import pandas as pd
from .serializers import ProductItemSerializer
from . import models
from celery import shared_task
from core.celery import BaseTaskWithRetry
from typing import Any, List

logger = logging.getLogger()

MAX_CHUNKS_SIZE = 100
FOREIGN_KEYS = [
    "item_group_id",
    "brand",
    "gender",
    "google_product_category",
    "product_type",
    "material",
    "pattern",
    "color",
]


class NullValueException(Exception):
    def __init__(self, message="Null value found in the cell"):
        self.message = message
        super().__init__(self.message)


class EmptyRowException(Exception):
    def __init__(self, message="Empty row found in the file"):
        self.message = message
        super().__init__(self.message)


def extract_header(file_path: str) -> list[str]:
    df = pd.read_excel(file_path, nrows=0)
    headers = df.columns.to_list()
    return headers


def end_task(start_time: float) -> None:
    end_time = time.time()
    models.Log.objects.create(
        message=end_time, status=models.LogStatus.INFO, remarks="END_TIME"
    )
    total_time = end_time - start_time
    models.Log.objects.create(
        message=total_time,
        status=models.LogStatus.INFO,
        remarks=f"TOTAL_TIME_{start_time}", 
    )
    logger.info(f"Total time taken: {total_time} seconds")


@shared_task(base=BaseTaskWithRetry)
def map_workbook_Json(
    file_path: str, start_row: int, stop_row: int, headers: list[str], start_time: float
) -> None:
    data = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=start_row + 1, max_row=stop_row):
            row_data = {}
            if all(cell.value is None for cell in row): 
                    raise EmptyRowException(f"Empty row found in the file with row index {row[0].row}") 
                
            for index, cell in enumerate(row[: len(headers)]):
                if cell.value is None:
                    e = NullValueException(
                        f"Null value found in the cell {cell.coordinate}"
                    )
                    logger.warning(e.message)
                    models.Log.objects.create(
                        message=e.__str__(),
                        status=models.LogStatus.WARNING,
                        remarks=f"WARNING_{start_time}",
                    )
                    cell.value = ""
                else:
                    cell.value = cell.value.strip()
                if headers[index] == "shipping(country:price)":
                    headers[index] = "shipping_cost"

                if headers[index] in FOREIGN_KEYS:
                    row_data[headers[index].lower()] = {"name": cell.value}
                else:
                    row_data[headers[index].lower()] = cell.value
            data.append(row_data)


    except EmptyRowException as e:
        logger.error(e.__str__())
        models.Log.objects.create(
            message=e.__str__(),
            status=models.LogStatus.ERROR,
            remarks=f"ERROR_{start_time}",
        )

    except Exception as e:
        logger.error(e.__str__())
        models.Log.objects.create(
            message=e.__str__(),
            status=models.LogStatus.ERROR,
            remarks=f"ERROR_{start_time}",
        )

    finally:
        serializer = ProductItemSerializer(data=data, many=True,context ={"start_time": start_time}) 
        serializer.is_valid(raise_exception=True)
        serializer.save()
        end_task(start_time)


@shared_task(base=BaseTaskWithRetry)
def serialize_and_save_json(filepath: str) -> None:
    start_time = time.time()
    models.Log.objects.create(
        message=start_time, status=models.LogStatus.INFO, remarks="START_TIME"
    )

    headers = extract_header(filepath)
    workbook = openpyxl.load_workbook(filepath)
    worksheets = workbook.active
    max_row = worksheets.max_row
    row_counter = max_row
    start_row = 1
    try:
        while row_counter > 0 or start_row < row_counter:
            if row_counter < MAX_CHUNKS_SIZE:
                if start_row > row_counter and row_counter < MAX_CHUNKS_SIZE:
                    end_row = max_row
                else:
                    end_row = row_counter
                map_workbook_Json.delay(
                    filepath, start_row, end_row, headers, start_time
                )
            else:
                end_row = MAX_CHUNKS_SIZE + start_row
                map_workbook_Json.delay(
                    filepath, start_row, end_row, headers, start_time
                )
                start_row = end_row
            row_counter = row_counter - MAX_CHUNKS_SIZE
    except Exception as e:
        logger.error(e.__str__())
        models.Log.objects.create(
            message=e.__str__(),
            status=models.LogStatus.ERROR,
            remarks=f"ERROR_{start_time}",
        )
        end_task(start_time)
